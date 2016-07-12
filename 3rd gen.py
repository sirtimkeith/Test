import sys
import os
import xlrd
from openpyxl import load_workbook

dataList = []
EXTENSION = (".xls")
#concatenation
def cat_man():
    catstring = ','.join(dataList)
    for f in catstring:
        wb = load_workbook(catstring)
        sheet = wb.get_sheet_by_name('Sheet1')
    '''
    try:
        for f in dataList:
            book = open_workbook(dataList,'r+')
            sheet = book.sheet_by_index(3)
            keys = [sheet.cell(0, col_index).value for col_index in xrange(sheet.ncols)]
            test_list = []
            for row_index in xrange(1, sheet.nrows):
                d = {keys[col_index]: sheet.cell(row_index, col_index).value
                     for col_index in xrange(sheet.ncols)}
                test_list.append(d)
        print (test_List)
        '''
    
#locating the files
def loc_man():
    while True:
        try:
            user_files = input("Enter the path of your file(s) or hit enter to see your current list. ")
            if user_files.endswith(EXTENSION) + os.path.exists(user_files):
                f = open(user_files,'r+')
                dataList.append(user_files)
                print("Found the file, adding it")
                f.close()
            elif user_files == '':
                print (dataList)
                continue_user = input("This is your file-list, press c to continue or b to add more files ")
                if continue_user == 'b':
                    loc_man()
                elif continue_user == 'c':
                    return cat_man()
            else:
                print("I either can't find it, or it's not a .xls file :(")
                loc_man()
        except OSError:
            print ("I either can't find it, or it's not a .xls file :(")
            loc_man()

loc_man()
cat_man()
